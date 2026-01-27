import sys
import json
import math
import shutil
import pandas as pd
import sqlalchemy
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
from utils import generate_filename, get_settings, minutes_to_readable, get_attendance_report_query, get_month_year

# --- Database Credentials ---
DB_HOST = ""
DB_USER = ""
DB_PASS = ""
DB_NAME = ""
engine = sqlalchemy.create_engine(f"mysql+pymysql://{DB_USER}:{DB_PASS}@{DB_HOST}/{DB_NAME}")


def main():

  if len(sys.argv) != 4:
    print("Usage: python script.py <emp_id> <start_date> <end_date>")
    sys.exit(1)

  # --- Attendance Details ---
  # emp_id = 152
  # start_date = '2025-01-01'
  # end_date = '2025-10-30'
  emp_id = sys.argv[1]
  start_date = sys.argv[2]
  end_date = sys.argv[3]


  # --- template source ---
  TEMPLATE_SOURCE = "template.xlsx"

  # ------ Color Settings ------
  setting = get_settings()
  colors = setting['colors']

  employees = []

  if emp_id == "0":
    EXCLUDED_ATTENDANCE_EMAILS = [
      'rashid@eurustechnologies.com',
      'azfar@eurustechnologies.com',
      'humayun@eurustechnologies.com',
      'bilal@eurustechnologies.com',
      'wajahat@eurustechnologies.com',
      'sajid@eurustechnologies.com',
      'babar.naseer@eurustechnologies.com',
      'raza@eurustechnologies.com',
      'waqar@eurustechnologies.com',
      'nabeel@eurustechnologies.com',
      'ali@eurustechnologies.com'
    ]
    query = 'SELECT emp_other_id, emp_work_email FROM hs_hr_employee WHERE emp_status IN (2,3);'

    df = pd.read_sql(query, engine)
    filtered_df = df[~df['emp_work_email'].isin(EXCLUDED_ATTENDANCE_EMAILS)]
    employees = filtered_df['emp_other_id'].tolist()
    # employees = [121, 116, 102, 129, 159, 152, 162, 168, 105, 164, 111, 132, 143, 147, 142, 157, 171, 100, 180, 175, 104, 188, 189, 136, 131, 149, 194, 185, 196, 117, 137, 187, 140, 144, 197, 153, 181, 134, 1, 200, 184, 110, 127, 192, 201, 203, 207, 209, 211]
  
  else:
    employees = [emp_id]
  
  print('Employees: ', employees)

  for emp in employees:

    print(f'⏳ Initiating attendance report for employee ID: {emp}...')

    query = get_attendance_report_query(str(emp), start_date, end_date, setting['margin_in_mins'])
    safe_query = query.replace('%', '%%')
    df = pd.read_sql(safe_query, engine)

    employee_name = ''.join(df['employee_name'][0].split(' '))
    selected_cols = ['attendance_date', 'shift', 'check_in_time', 'status', 'late_in_mins']
    df_selected = df[selected_cols]

    output_filename = generate_filename(employee_name, start_date, end_date)
    shutil.copy(TEMPLATE_SOURCE, output_filename)

    wb = load_workbook(output_filename)
    ws = wb.active

    ##############################
    # Columns
    ##############################
    current_month = None
    start_row = 14
    write_row = start_row

    for row in dataframe_to_rows(df_selected, index=False, header=False):
      month_year = get_month_year(row[0])

      # Insert month heading before the first entry of each month
      if month_year != current_month:
        if current_month != None:
          write_row += 1

        ws.merge_cells(start_row=write_row, start_column=1, end_row=write_row, end_column=5)
        month_cell = ws.cell(row=write_row, column=1, value=month_year)
        month_cell.fill = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
        month_cell.font = Font(color='ffffff', bold=True, size=12)
        month_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[write_row].height = 23

        current_month = month_year
        write_row += 1


      ws.row_dimensions[write_row].height = 18
      status = row[3]
      color_category = 'ALL_LEAVES' if '_LEAVE' in status else status

      for c_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=write_row, column=c_idx, value=value)

        # Status formatting
        if c_idx == 4:
          cell.value = value.replace('_', ' ').title()
        
        # late column
        elif c_idx == 5:
          if pd.isna(value):
            is_absent = True
          else:
            late_duration = minutes_to_readable(value)
            is_absent = pd.isna(float(late_duration.split(' ')[0]))
          
          cell.value = '' if is_absent else late_duration

        cell.alignment = Alignment(horizontal='left')
        cell.fill = PatternFill(start_color=colors[color_category]['bg'], end_color=colors[color_category]['bg'], fill_type='solid')
        cell.font = Font(color=colors[color_category]['text'], bold=False, size=11)

      write_row += 1

    ##############################
    # Attendance Summary
    ##############################
    total_public_holidays = df_selected['status'].eq('PUBLIC_HOLIDAY').sum()
    ws['E6'] = total_public_holidays
    ws['E6'].font = Font(bold=True, size=11)
    ws['E6'].alignment = Alignment(horizontal='left')

    total_working_days = len(df_selected) - total_public_holidays
    ws['E3'] = total_working_days
    ws['E3'].font = Font(bold=True, size=11)
    ws['E3'].alignment = Alignment(horizontal='left')

    total_late_arrivals = df_selected['status'].eq('LATE').sum()
    ws['B6'] = total_late_arrivals
    ws['B6'].font = Font(bold=True, size=11)
    ws['B6'].alignment = Alignment(horizontal='left')

    # Weighted WFH count (0.5 for half-day, 1 for full)
    total_wfh = (
      1 * df_selected['status'].eq('WFH').sum() +
      0.5 * df_selected['status'].eq('HALF_DAY_WFH').sum() +
      0.5 * df_selected['status'].eq('HALF_DAY_LEAVE_AND_WFH').sum()
    )
    ws['B9'] = total_wfh
    ws['B9'].font = Font(bold=True, size=11)
    ws['B9'].alignment = Alignment(horizontal='left')

    total_leaves_not_applied = df_selected['status'].eq('LEAVE_NOT_APPLIED').sum()
    ws['E9'] = total_leaves_not_applied
    ws['E9'].font = Font(bold=True, size=11)
    ws['E9'].alignment = Alignment(horizontal='left')


    ##############################
    # Performance Calculation
    ##############################
    # Penalty system
    penalty_per_minute = 0.015  # (21.6 minutes)
    present_status_weights = {
      'PRESENT': 1,
      'LATE': 1,
      'WFH': 1,
      'HALF_DAY_LEAVE': 0.5,
      'HALF_DAY_LEAVE_AND_WFH': 0.5
    }

    total_present_days = (df_selected['status'].map(present_status_weights).fillna(0)).sum()
    total_late_minutes = df_selected.loc[df_selected['late_in_mins'] > 10, 'late_in_mins'].fillna(0).sum()
    effective_days = total_present_days - (total_late_minutes * penalty_per_minute)
    performance_percent = round((effective_days / total_working_days) * 100, 2)
    ws['B3'] = f'{performance_percent}%'
    ws['B3'].font = Font(bold=True, size=11)
    ws['B3'].alignment = Alignment(horizontal='left')

    # Save updated file
    wb.save(output_filename)
    print(f"📋 Attendance report exported as {output_filename}")
    print('📊 Attendance Summary:')
    print(f'Performance         : {performance_percent}')
    print(f'Days Present        : {total_present_days}/{total_working_days}')
    print(f'Late Arrivals       : {total_late_arrivals}')
    print(f'Public Holidays     : {total_public_holidays}')
    print(f'Work From Home      : {total_wfh}')
    print(f'Leaves Not Applied  : {total_leaves_not_applied}')



if __name__ == "__main__":
  main()
