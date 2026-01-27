from utils import get_settings
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

wb = Workbook()
ws = wb.active
# ws.title = "Employee Attendance"
ws.sheet_view.showGridLines = False

# ------ Color Settings ------
colors = get_settings()['colors']

# --- Set column widths ---
column_widths = {
  'A': 23,
  'B': 13,
  'C': 18,
  'D': 19,
  'E': 12
}
for col, width in column_widths.items():
  ws.column_dimensions[col].width = width


# --- Set header section row heights ---
row_heights = {
  1: 28,  # Header row
  2: 16,
  3: 24,
  4: 18,
  5: 15,
  6: 24,
  7: 18,
  8: 14,
  9: 24,
  10: 18,
  11: 14,
  12: 14,
  13: 26
}
for row, height in row_heights.items():
  ws.row_dimensions[row].height = height


# Heading
ws.merge_cells('A1:E1')
ws['A1'] = 'ATTENDANCE REPORT'
ws['A1'].fill = PatternFill(start_color='1f2937', end_color='1f2937', fill_type='solid')
ws['A1'].font = Font(color='FFFFFF', bold=True, size=14)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')


##############################
# Widgets
##############################
# --- Setting Attendance Performance Widget ---
for row in ws['A3:B4']:
  for cell in row:
    cell.fill = PatternFill(start_color='eef2ff', end_color='eef2ff', fill_type='solid')
ws['A3'] = 'Attendance Rate'
ws['A3'].font = Font(bold=True, size=11)
ws['A4'] = 'Overall performance'
ws['A4'].font = Font(color='6b7280', bold=False, size=10)

# --- Setting Total Days Widget ---
for row in ws['D3:E4']:
  for cell in row:
    cell.fill = PatternFill(start_color='eef2ff', end_color='eef2ff', fill_type='solid')
ws['D3'] = 'Total Days'
ws['D3'].font = Font(bold=True, size=11)
ws['D4'] = 'Working days'
ws['D4'].font = Font(color='6b7280', bold=False, size=10)

# --- Setting Late Arrivals Widget ---
for row in ws['A6:B7']:
  for cell in row:
    cell.fill = PatternFill(start_color=colors['LATE']['bg'], end_color=colors['LATE']['bg'], fill_type='solid')
ws['A6'] = 'Late Arrivals'
ws['A6'].font = Font(bold=True, size=11)
ws['A7'] = 'Days late'
ws['A7'].font = Font(color='6b7280', bold=False, size=10)

# --- Setting Public Holidays Widget ---
for row in ws['D6:E7']:
  for cell in row:
    cell.fill = PatternFill(start_color=colors['PUBLIC_HOLIDAY']['bg'], end_color=colors['PUBLIC_HOLIDAY']['bg'], fill_type='solid')
ws['D6'] = 'Public Holidays'
ws['D6'].font = Font(bold=True, size=11)
ws['D7'] = 'Days off'
ws['D7'].font = Font(color='6b7280', bold=False, size=10)

# --- Setting WFH Widget ---
for row in ws['A9:B10']:
  for cell in row:
    cell.fill = PatternFill(start_color=colors['WFH']['bg'], end_color=colors['WFH']['bg'], fill_type='solid')
ws['A9'] = 'Work From Home'
ws['A9'].font = Font(bold=True, size=11)
ws['A10'] = 'WFH'
ws['A10'].font = Font(color='6b7280', bold=False, size=10)

# --- Setting Leaves Not Applied Widget ---
for row in ws['D9:E10']:
  for cell in row:
    cell.fill = PatternFill(start_color=colors['LEAVE_NOT_APPLIED']['bg'], end_color=colors['LEAVE_NOT_APPLIED']['bg'], fill_type='solid')
ws['D9'] = 'Leaves Not Applied'
ws['D9'].font = Font(bold=True, size=11)
ws['D10'] = 'Unmarked leaves'
ws['D10'].font = Font(color='6b7280', bold=False, size=10)


##############################
# Columns
##############################
columns = ['Date', 'Shift Time', 'Check-in Time', 'Status', 'Late']
for col_num, column_title in enumerate(columns, start=1):
  cell = ws.cell(row=13, column=col_num, value=column_title)
  cell.fill = PatternFill(start_color='1f2937', end_color='1f2937', fill_type='solid')
  cell.font = Font(color='FFFFFF', bold=True, size=12)
  cell.alignment = Alignment(horizontal='center', vertical='center')


# Save workbook
wb.save('template.xlsx')

print("✅ Excel file 'template.xlsx' created successfully.")